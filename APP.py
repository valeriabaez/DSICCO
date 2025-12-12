# app.py
import streamlit as st
import pandas as pd
import os
from io import BytesIO

# ---------------------------------------------------------
# CONFIGURACIÓN
# ---------------------------------------------------------
st.set_page_config(page_title="DSICCO – Resúmenes A 2025", layout="wide", page_icon="escudo.png")

st.markdown("""
<div style='text-align:center; background-color:#003366; padding:15px; border-radius:10px;'>
    <h1 style='color:white;'>🛡️ DSICCO – Carga y Resúmenes 2025</h1>
    <p style='color:white;'>Subí tu archivo DSICCO.xlsx con las hojas ALLANAMIENTOS y ARMAS</p>
</div>
""", unsafe_allow_html=True)


# ---------------------------------------------------------
# CARPETA UPLOADS
# ---------------------------------------------------------
UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
SAVED_FILE = os.path.join(UPLOAD_FOLDER, "DSICCO.xlsx")

# ---------------------------------------------------------
# FUNCIONES AUXILIARES
# ---------------------------------------------------------
def nombre_mes(num):
    meses = [
        "SIN MES","ENERO","FEBRERO","MARZO","ABRIL","MAYO","JUNIO",
        "JULIO","AGOSTO","SEPTIEMBRE","OCTUBRE","NOVIEMBRE","DICIEMBRE"
    ]
    try:
        n = int(num)
        return meses[n] if 1 <= n <= 12 else "SIN MES"
    except:
        return "SIN MES"

def cargar_excel(path):
    return pd.read_excel(path, sheet_name=None)


def build_blocks(df, mes_col, mes_name_col, unidad_col="UNIDAD", interv_col="INTERVENCION", cant_col="CANTIDAD"):
    blocks = []
    total_general = 0
    for mes in sorted(df[mes_col].unique()):
        df_mes = df[df[mes_col] == mes]
        if df_mes.empty:
            continue

        mes_label = df_mes[mes_name_col].iloc[0]
        blocks.append([mes_label, "", "", ""])

        for _, r in df_mes.iterrows():
            blocks.append([
                "",
                r.get(unidad_col, ""),
                r.get(interv_col, "ALLANAMIENTO") if interv_col else "ALLANAMIENTO",
                int(r.get(cant_col, 0))
            ])

        subtotal = int(df_mes[cant_col].sum())
        blocks.append(["Subtotal", "", "", subtotal])
        total_general += subtotal

    blocks.append(["TOTAL GENERAL", "", "", total_general])
    return blocks


# ---------------------------------------------------------
# EXPORTACIÓN A EXCEL CON CELDAS COMBINADAS
# ---------------------------------------------------------
def export_excel(blocks_allan, blocks_armas):
    import openpyxl

    output = BytesIO()

    df_allan = pd.DataFrame(blocks_allan, columns=["Mes", "Unidad", "Intervención", "Cantidad"])
    df_armas = pd.DataFrame(blocks_armas, columns=["Mes", "Unidad", "Intervención", "Cantidad"])

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_allan.to_excel(writer, sheet_name="ALLANAMIENTOS", index=False)
        df_armas.to_excel(writer, sheet_name="ARMAS", index=False)

        # --------------------- ALLANAMIENTOS ------------------------
        ws = writer.book["ALLANAMIENTOS"]
        merge_start = None
        last_mes = None

        for row in range(2, ws.max_row + 1):
            val = ws[f"A{row}"].value

            if val not in ["", None, "Subtotal", "TOTAL GENERAL"]:
                if merge_start is not None:
                    ws.merge_cells(f"A{merge_start}:A{row-1}")
                merge_start = row
                last_mes = val

            if val == "Subtotal":
                if merge_start:
                    ws.merge_cells(f"A{merge_start}:A{row-1}")
                merge_start = None
                last_mes = None

        if merge_start:
            ws.merge_cells(f"A{merge_start}:A{ws.max_row}")

        # --------------------- ARMAS ------------------------
        ws2 = writer.book["ARMAS"]
        merge_start = None
        last_mes = None

        for row in range(2, ws2.max_row + 1):
            val = ws2[f"A{row}"].value

            if val not in ["", None, "Subtotal", "TOTAL GENERAL"]:
                if merge_start:
                    ws2.merge_cells(f"A{merge_start}:A{row-1}")
                merge_start = row
                last_mes = val

            if val == "Subtotal":
                if merge_start:
                    ws2.merge_cells(f"A{merge_start}:A{row-1}")
                merge_start = None
                last_mes = None

        if merge_start:
            ws2.merge_cells(f"A{merge_start}:A{ws2.max_row}")

    return output.getvalue()


# ---------------------------------------------------------
# SUBIR ARCHIVO NUEVO
# ---------------------------------------------------------
uploaded = st.file_uploader("📂 Seleccioná archivo Excel", type=["xlsx"])

if uploaded:
    with open(SAVED_FILE, "wb") as f:
        f.write(uploaded.getvalue())
    st.success("✔ Archivo cargado y reemplazado correctamente.")

# ---------------------------------------------------------
# CARGAR ARCHIVO EXISTENTE
# ---------------------------------------------------------
if not os.path.exists(SAVED_FILE):
    st.warning("📁 Todavía no hay archivo cargado.")
    st.stop()

try:
    excel = cargar_excel(SAVED_FILE)
except Exception as e:
    st.error(f"❌ Error al abrir el archivo guardado: {e}")
    st.stop()

if "ALLANAMIENTOS" not in excel or "ARMAS" not in excel:
    st.error("❌ El archivo debe contener ALLANAMIENTOS y ARMAS.")
    st.stop()

allan = excel["ALLANAMIENTOS"].copy()
allan.columns = allan.columns.str.upper().str.strip()

armas = excel["ARMAS"].copy()
armas.columns = armas.columns.str.upper().str.strip()

# ---------------------------------------------------------
# PROCESAR ALLANAMIENTOS
# ---------------------------------------------------------
st.markdown("## 🔵 ALLANAMIENTOS")

if "FECHA" not in allan.columns:
    st.error("❌ ALLANAMIENTOS debe tener FECHA.")
    st.stop()

allan["FECHA"] = pd.to_datetime(allan["FECHA"], errors="coerce")
allan["MES"] = allan["FECHA"].dt.month.fillna(0).astype(int)
allan["MES_NOMBRE"] = allan["MES"].apply(nombre_mes)
allan["POSITIVO_FLAG"] = allan["RESULTADO"].astype(str).str.upper().str.contains("POS", na=False)
allan["NEGATIVO_FLAG"] = allan["RESULTADO"].astype(str).str.upper().str.contains("NEG", na=False)
allan["CANTIDAD"] = 1

resumen_allan = (
    allan.groupby(["MES", "MES_NOMBRE", "UNIDAD"], as_index=False)
    .agg({"POSITIVO_FLAG": "sum", "NEGATIVO_FLAG": "sum", "CANTIDAD": "sum"})
)

blocks_allan = build_blocks(
    resumen_allan,
    "MES",
    "MES_NOMBRE",
    unidad_col="UNIDAD",
    interv_col=None,
    cant_col="CANTIDAD"
)

for mes in sorted(resumen_allan["MES"].unique()):
    df_mes = resumen_allan[resumen_allan["MES"] == mes]

    with st.expander(f"📅 {df_mes['MES_NOMBRE'].iloc[0]}"):

        # 🔥 Solo mostramos UNIDAD – POS – NEG – TOTAL (sin repetir MES)
        tabla = df_mes[["UNIDAD", "POSITIVO_FLAG", "NEGATIVO_FLAG", "CANTIDAD"]].rename(columns={
            "UNIDAD": "Unidad",
            "POSITIVO_FLAG": "Positivos",
            "NEGATIVO_FLAG": "Negativos",
            "CANTIDAD": "Total"
        })

        st.table(tabla)

        st.markdown(f"**Subtotal:** {df_mes['CANTIDAD'].sum()}")

# ---------------------------------------------------------
# TOTALES DE ALLANAMIENTOS (debajo de los expanders)
# ---------------------------------------------------------

total_positivos = int(allan["POSITIVO_FLAG"].sum())
total_negativos = int(allan["NEGATIVO_FLAG"].sum())
total_allanamientos = int(allan["CANTIDAD"].sum())

st.write("### Total Positivos")
st.markdown(f"<h2 style='margin-top:-10px;'>{total_positivos}</h2>", unsafe_allow_html=True)

st.write("### Total Negativos")
st.markdown(f"<h2 style='margin-top:-10px;'>{total_negativos}</h2>", unsafe_allow_html=True)

st.write("### TOTAL Allanamientos")
st.markdown(f"<h2 style='margin-top:-10px;'>{total_allanamientos}</h2>", unsafe_allow_html=True)

# ---------------------------------------------------------
# ARMAS
# ---------------------------------------------------------
st.markdown("## 🔴 ARMAS")

required = ["FECHA", "TIPO", "INTERVENCION", "CANTIDAD"]
for col in required:
    if col not in armas.columns:
        st.error(f"❌ La hoja ARMAS debe tener {col}.")
        st.stop()

armas["FECHA"] = pd.to_datetime(armas["FECHA"], errors="coerce")
armas["MES"] = armas["FECHA"].dt.month.fillna(0).astype(int)
armas["MES_NOMBRE"] = armas["MES"].apply(nombre_mes)

armas_validas = armas[
    armas["TIPO"].astype(str).str.upper().str.contains("ARMA|TUMBERA", regex=True, na=False)
].copy()

armas_validas["CANTIDAD"] = (
    pd.to_numeric(armas_validas["CANTIDAD"], errors="coerce")
    .fillna(1)
    .astype(int)
)

resumen_armas = (
    armas_validas.groupby(["MES", "MES_NOMBRE", "UNIDAD", "INTERVENCION"], as_index=False)
    .agg({"CANTIDAD": "sum"})
)

blocks_armas = build_blocks(
    resumen_armas,
    "MES",
    "MES_NOMBRE",
    unidad_col="UNIDAD",
    interv_col="INTERVENCION",
    cant_col="CANTIDAD"
)

for mes in sorted(resumen_armas["MES"].unique()):
    df_mes = resumen_armas[resumen_armas["MES"] == mes]
    with st.expander(f"📅 {df_mes['MES_NOMBRE'].iloc[0]}"):
        st.table(df_mes)

st.metric("Total armas", int(resumen_armas["CANTIDAD"].sum()))

# ---------------------------------------------------------
# RESUMEN RÁPIDO
# ---------------------------------------------------------
st.markdown("## 📊 Resumen rápido")

col1, col2 = st.columns(2)

with col1:
    st.markdown("**Armas por mes (ordenado):**")
    total_armas_mes = (
        resumen_armas.groupby(["MES", "MES_NOMBRE"], as_index=False)["CANTIDAD"]
        .sum()
        .sort_values("MES")
    )
    st.table(total_armas_mes[["MES_NOMBRE", "CANTIDAD"]])

with col2:
    st.markdown("**Armas por procedimiento:**")
    st.table(
        resumen_armas.groupby("INTERVENCION")["CANTIDAD"]
        .sum()
        .reset_index()
    )

# ---------------------------------------------------------
# DESCARGA EXCEL
# ---------------------------------------------------------
excel_bytes = export_excel(blocks_allan, blocks_armas)

st.download_button(
    label="📥 Descargar Resúmenes en EXCEL",
    data=excel_bytes,
    file_name="Resumenes_DSICCO.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
